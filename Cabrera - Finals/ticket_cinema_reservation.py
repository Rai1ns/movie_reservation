from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import ImageTk, Image
import time
import datetime as dt

root = tk.Tk()
root.geometry("1312x650")
root.title("Welcome To CiNEMA")
root.config(bg="Black")
root.resizable(0, 0)
messagebox.showinfo("REMINDER!", "1 person, 1 ticket only thankyou!!")

file = "ticket_reservation.xlsx"

excel_con = Workbook()
excel_con = load_workbook(file)
excel_activate = excel_con.active

def reserved_btn():
    newroot = tk.Toplevel()
    newroot.geometry("700x310")
    newroot.title("Data Reserved")
    newroot.config(bg="#8B0000")
    newroot.resizable(0, 0)
    
    excel_con = Workbook()
    excel_con = load_workbook(file)
    excel_activate = excel_con.active
    
    def home_btn():
        newroot.destroy()
        
    def del_btn():
        root_del = tk.Toplevel()
        root_del.geometry("250x200")
        root_del.resizable(0, 0)
        root_del.config(bg="#8B0000")
        newroot.destroy()
        
        def deltn():
            for each_cell in range(2, (excel_activate.max_row)+1):
                if (del_ur_name.get() ==  excel_activate['A'+str(each_cell)].value):
                    Found = True
                    cell_address = each_cell            
                    break;
                else:
                    Found=False
            if (Found == True):
                msg = messagebox.askquestion("Are you sure?", 'Are you sure you to delete this data named "' + del_ur_name.get() + '"?')
                if msg == "yes":    
                    excel_activate.delete_rows(cell_address)
                    messagebox.showinfo("INFO",'Data Deleted named "' + del_ur_name.get() + '"')
                else:
                    pass
                excel_con.save(file)
                cans_btn()
            elif (Found == False):
                messagebox.showerror("NOT FOUND", 'Data named"' + del_ur_name.get() + '" not found')
                cans_btn()
                
        def cans_btn():
            root_del.destroy()
            newroot.destroy()
        
        def edit_btn():
            name_row = 0
            for each_cell in range(2, (excel_activate.max_row)+1):
                if (del_ur_name.get() ==  excel_activate['A'+str(each_cell)].value):
                    Found = True;
                    name_row = each_cell
                    break;
                else:
                    Found=False
            if Found == False:
                messagebox.showinfo("Not Found", '"' + del_ur_name.get() + '" Not Found')
                cans_btn()
                           
            if Found == True:
                edit_interface = tk.Toplevel()
                edit_interface.geometry('500x400')
                edit_interface.title('Edit Data')
                edit_interface.config(bg="#8B0000")
                cans_btn()
                
                def old_name():
                    if name_var.get()==1:
                        name = excel_activate['A'+str(name_row)].value
                        nameTxt.insert(0, excel_activate['A'+str(name_row)].value)
                        nameTxt.config(state="disabled")
                    elif name_var.get() ==0:
                        nameTxt.config(state="normal")
                        nameTxt.delete(0, END)
                def old_age():
                    if ageChoice.get()==1:
                        age = excel_activate['B'+str(name_row)].value
                        age_edit2.insert(0, age)
                        age_edit2.config(state="disabled")
                    elif ageChoice.get()==0:
                        age_edit2.config(state="normal")
                        age_edit2.delete(0, END)
                        
                def old_movie():
                    if movieChoise.get()==1:
                        movie = excel_activate['C'+str(name_row)].value
                        movie_edit2.insert(0, movie)
                        movie_edit2.config(state="disabled")
                    elif movieChoise.get() ==0:
                        movie_edit2.config(state="normal")
                        movie_edit2.delete(0, END)
                        
                def old_time():
                    if timeChoise.get()==1:
                        time = excel_activate['D'+str(name_row)].value
                        timeExcel.set(time)
                        time1_["state"] = "disable"
                        time2_["state"] = "disable"
                        time3_["state"] = "disable"
                        time4_["state"] = "disable"
                    elif timeChoise.get()==0:
                        time1_["state"] = "normal"
                        time2_["state"] = "normal"
                        time3_["state"] = "normal"
                        time4_["state"] = "normal"
                        timeExcel.set("")
                        
                def old_seat():
                    if seatChoise.get() == 1:
                        seat = excel_activate['E'+str(name_row)].value
                        seat_combo_a.insert(0, seat)
                        seat_combo_a.config(state="disabled")
                        seat_combo_b.config(state="disabled")
                        seat_combo_c.config(state="disabled")
                        seat_combo_d.config(state="disabled")
                        seat_combo_e.config(state="disabled")
                        seat_combo_f.config(state="disabled")
                    elif seatChoise.get() == 0:
                        seat_combo_a.config(state="normal")
                        seat_combo_b.config(state="normal")
                        seat_combo_c.config(state="normal")
                        seat_combo_d.config(state="normal")
                        seat_combo_e.config(state="normal")
                        seat_combo_f.config(state="normal")
                        seat_combo_a.delete(0, END)
                        
                        
                def save_btn():
                    if nameTxt.get() == "":
                        name = excel_activate['A'+str(name_row)].value
                    else:
                        name = nameTxt.get()
                    if age_edit2.get() == "":
                        age = excel_activate['B'+str(name_row)].value
                    else:
                        age = age_edit2.get()
                    if movieExcel.get() == "":
                        movie = excel_activate['C'+str(name_row)].value
                    else:
                        movie = movieExcel.get()
                    if timeExcel.get() == "":
                        time = excel_activate['D'+str(name_row)].value
                    else:
                        time = timeExcel.get()
                    for each_cell in range(2, (excel_activate.max_row)+1):
                        if (seatExcel.get() ==  excel_activate['E'+str(each_cell)].value):
                            Found = True         
                            break;
                        else:
                            Found=False
                    if Found == True:
                        messagebox.showerror("Occupied", "Seat is Already Occupied")
                    else:
                        seat = seatExcel.get()
                        
                        excel_activate['A'+str(name_row)] = name
                        excel_activate['B'+str(name_row)] = age
                        excel_activate['C'+str(name_row)] = movie
                        excel_activate['D'+str(name_row)] = time
                        excel_activate['E'+str(name_row)] = seat


                        excel_con.save(file)
                        messagebox.showinfo("UPDATED","DATA HAS BEEN UPDATED")
                
                
                ageExcel = StringVar()
                movieExcel = StringVar()
                timeExcel = StringVar()
                seatExcel = StringVar()

                EditLabel = Label(edit_interface, text="Edit Form ", font=("Arial Black", 15), bg="#8B0000")
                EditLabel.grid(row=0, column=1, pady=5)

                nameLbl=LabelFrame(edit_interface,text="Name", font=("Arial Black", 10), bg="#8B0000")
                nameLbl.grid(row=1, column=0)
                

                nameTxt=Entry(nameLbl, width=10, font=("Arial Black", 10))
                nameTxt.grid(row=2, column=0)

                name_var = IntVar()
                nameChk = Checkbutton(edit_interface,text="Previous", variable=name_var, command=old_name, font=("Arial Black", 10), bg="#8B0000")
                nameChk.grid(row=3, column=0)
                
                age_edit=LabelFrame(edit_interface,text="Age", font=("Arial Black", 10), bg="#8B0000")
                age_edit.grid(row=1, column=1, pady=5)

                age_edit_list = []
                for i in range(18,81):
                    age_edit_list.append(i)
                    
                age_edit2=ttk.Combobox(age_edit,width=10, font=("Arial Black", 10), textvariable=ageExcel, values=age_edit_list)
                age_edit2.grid(row=0, column=0)

                ageChoice = IntVar()
                ageChk = Checkbutton(edit_interface, text="Previous", font=("Arial Black", 10), bg="#8B0000", variable=ageChoice,command=old_age, state="normal")
                ageChk.grid(row=3, column=1)

                movie_edit =LabelFrame(edit_interface,text="Movie", font=("Arial Black", 10), bg="#8B0000")
                movie_edit.grid(row=1, column=3, pady=5)

                movie_edit_list = ["Avatar", "Avengers", "Black Adam", "Chappie", "Fast X", "TransFormer"]
                movie_edit2=ttk.Combobox(movie_edit,width=10, font=("Arial Black", 10), textvariable=movieExcel, values=movie_edit_list)
                movie_edit2.grid(row=0, column=0)

                movieChoise = IntVar()
                movieChk = Checkbutton(edit_interface, text="Previous", variable=movieChoise, font=("Arial Black", 10), bg="#8B0000",command=old_movie)
                movieChk.grid(row=3, column=3)

                time_edit = LabelFrame(edit_interface, text="Time", font=("Arial Black", 10), bg="#8B0000")
                time_edit.grid(row=4, column=0, pady=5, padx=5)
                
                time_edit_list = ["8:00AM - 10:30AM", "11:00AM - 1:30PM", "2:00PM - 4:30PM", "5:00PM - 7:30PM"]
                time1_ = Radiobutton(time_edit, text=time_edit_list[0], variable=timeExcel, value=time_edit_list[0], font=("Arial Black", 10), bg="#8B0000")
                time1_.grid(row=4, column=0)

                time2_ = Radiobutton(time_edit, text=time_edit_list[1], variable=timeExcel, value=time_edit_list[1], font=("Arial Black", 10), bg="#8B0000")
                time2_.grid(row=5, column=0)

                time3_ = Radiobutton(time_edit, text=time_edit_list[2], variable=timeExcel, value=time_edit_list[2], font=("Arial Black", 10), bg="#8B0000")
                time3_.grid(row=6, column=0)

                time4_ = Radiobutton(time_edit, text=time_edit_list[3], variable=timeExcel, value=time_edit_list[3], font=("Arial Black", 10), bg="#8B0000")
                time4_.grid(row=7, column=0)
                
                timeChoise = IntVar()
                timeChk = Checkbutton(edit_interface, text="Previous", font=("Arial Black", 10), bg="#8B0000", variable=timeChoise, command=old_time)
                timeChk.grid(row=5, column=0)
                
                seat_edit = LabelFrame(edit_interface, text="Seat", font=("Arial Black", 10), bg="#8B0000")
                seat_edit.grid(row=4, column=1, pady=5, columnspan=3)
                
                edit_seat_list_a = []
                for a in range(1,12):
                    edit_seat_list_a.append("A"+str(a))
                    
                edit_seat_list_b = []
                for b in range(1,12):
                    edit_seat_list_b.append("B"+str(b))
                    
                edit_seat_list_c = []
                for c in range(1,12):
                    edit_seat_list_c.append("C"+str(c))
                    
                edit_seat_list_d = []
                for d in range(1,12):
                    edit_seat_list_d.append("D"+str(d))
                    
                edit_seat_list_e = []
                for e in range(1,12):
                    edit_seat_list_e.append("E"+str(e))
                    
                edit_seat_list_f = []
                for f in range(1,12):
                    edit_seat_list_f.append("F"+str(f))
                    
                seat_combo_a = ttk.Combobox(seat_edit, font=("Arial Black", 10),width=3, values=edit_seat_list_a, textvariable=seatExcel)
                seat_combo_a.grid(row=0, column=0,pady=10, padx=10)

                seat_combo_b = ttk.Combobox(seat_edit, font=("Arial Black", 10),width=3, values=edit_seat_list_b, textvariable=seatExcel)
                seat_combo_b.grid(row=0, column=1,pady=10, padx=10)

                seat_combo_c = ttk.Combobox(seat_edit, font=("Arial Black", 10),width=3, values=edit_seat_list_c, textvariable=seatExcel)
                seat_combo_c.grid(row=0, column=2,pady=10, padx=10)

                seat_combo_d = ttk.Combobox(seat_edit, font=("Arial Black", 10),width=3, values=edit_seat_list_d, textvariable=seatExcel)
                seat_combo_d.grid(row=1, column=0,pady=10, padx=10)

                seat_combo_e = ttk.Combobox(seat_edit, font=("Arial Black", 10),width=3, values=edit_seat_list_e, textvariable=seatExcel)
                seat_combo_e.grid(row=1, column=1,pady=10, padx=10)

                seat_combo_f = ttk.Combobox(seat_edit, font=("Arial Black", 10),width=3, values=edit_seat_list_f, textvariable=seatExcel)
                seat_combo_f.grid(row=1, column=2,pady=10)
                
                seatChoise = IntVar()
                seatChk = Checkbutton(edit_interface, text="Previous", font=("Arial Black", 10), bg="#8B0000", variable=seatChoise, command=old_seat)
                seatChk.grid(row=5, column=1, padx=10, pady=10, sticky=E)
                
                save = Button(edit_interface, font=("Arial Black", 15), text="Save",command=save_btn)
                save.grid(row=11, column=1)
                
                
                edit_interface.mainloop()
            
        
        del_name = Label(root_del, text="Name", font=("Arial Black", 20), bg="#8B0000")
        del_name.grid(row=0, column=0, pady=5, columnspan=3, padx=60)
        
        del_ur_name = Entry(root_del, font=("Arial Black", 12))
        del_ur_name.grid(row=1, column=0, columnspan=3, pady=5, padx=12)
        
        btn_del = Button(root_del, text="Delete", font=("Arial Black", 12), command=lambda:deltn())
        btn_del.grid(row=2, column=2, pady=5)
        
        btn_edit = Button(root_del, text="Edit", font=("Arial Black", 12), width=5, command=lambda:edit_btn())
        btn_edit.grid(row=2, column=0, pady=5)
        
        can_btn = Button(root_del, text="Cancel", font=("Arial Black", 12), command=lambda:cans_btn())
        can_btn.grid(row=3, column=0, pady=5, columnspan=3, padx=60)
        
        
        root_del.mainloop()
         
    treev = ttk.Treeview(newroot)
    treescrolly = Scrollbar(newroot, orient="vertical", command=treev.yview)
    treev.configure(yscrollcommand=treescrolly.set)
    treescrolly.pack(side =RIGHT,fill=Y)  

    treev['columns'] = ("Name", "Age", "Movies", "Time", "Seat")
    treev.column("#0", width=120, minwidth=25)
    treev.column("Name", anchor=W, width=120)
    treev.column("Age", anchor=W, width=120)
    treev.column("Movies",  anchor=CENTER, width=120)
    treev.column("Time", anchor=W, width=120)
    treev.column("Seat", anchor=W, width=70)

    treev.heading("#0", text="No.", anchor=W)
    treev.heading("Name", text="Name", anchor=W)
    treev.heading("Age", text="Age", anchor=W)
    treev.heading("Movies", text="Movies", anchor=CENTER)
    treev.heading("Time", text="Time", anchor=W)
    treev.heading("Seat", text="Seat", anchor=W)

    for each_cell in range(2, (excel_activate.max_row)+1):
        treev.insert(parent='', index="end", text=str(each_cell),values=(excel_activate['A'+str(each_cell)].value,\
                                                                        excel_activate['B'+str(each_cell)].value, \
                                                                        excel_activate['C'+str(each_cell)].value, \
                                                                        excel_activate['D'+str(each_cell)].value, \
                                                                        excel_activate['E'+str(each_cell)].value))
    treev.pack(pady=10)
    
    dne = Button(newroot, text="Edit/Delete Data", font=("Arial Black", 15), command=lambda:del_btn())
    dne.pack(side=LEFT, padx=10)
    
    home = Button(newroot, text="Home", font=("Arial Black", 15), command=lambda:home_btn())
    home.pack(side=RIGHT)
    
    
    
    newroot.mainloop()
    
def exit_btn_():
    message = messagebox.askquestion("EXIT", "Are you sure you want to exit?", icon="warning")
    if message == 'yes':
        root.destroy()
    
def done_btn():
    excel_con = Workbook()
    excel_con = load_workbook(file)
    excel_activate = excel_con.active
    
    name = ur_name.get()
    age = age_var.get()
    movie = ur_mov_var.get()
    seat = seat_var.get()
    time = time_var.get()
    
    if name == "" or age == "" or movie == "" or seat == "" or time == "":
        messagebox.showerror("Error", "Please Fill out the FORM!!")
    else:
        latest_row = str(excel_activate.max_row + 1)
        Found = False
        for each_cell in range(2, excel_activate.max_row+1):
            if name == excel_activate["A"+str(each_cell)].value or seat == excel_activate["E"+str(each_cell)].value:
                Found = True
                break
        if Found == True:
            messagebox.showerror("DATA", "Name is already exist or Seat Is already Occupied, Please Change it.")
        else:
            excel_activate["A"+latest_row] = name
            excel_activate["B"+latest_row] = age
            excel_activate["C"+latest_row] = movie
            excel_activate["D"+latest_row] = time
            excel_activate["E"+latest_row] = seat
            messagebox.showinfo("SUCCESS", '"' + ur_name.get() + '" Successfully Added')
            excel_con.save(file)
            
            name = ur_name.delete(0, END)
            age = age_var.set("")
            movie = ur_mov_var.set("")
            seat = seat_var.set("")
            time = time_var.set("")
            
def update_clock():
    current_time = time.strftime("%I:%M:%S %p")
    time_label.config(text=current_time)
    root.after(1000, update_clock)    
    
frame_left = Frame(root, bg="#8B0000", height=600, width=500)
frame_left.grid(row=0, column=0, padx=10)

frame_right = Frame(root, bg="Black")
frame_right.grid(row=0, column=1, padx=10)    
    
            
cin = Label(frame_left, text="Ticket 4\n CiNEMA", font=("Stencil", 30), bg="#8B0000")
cin.grid(row=0, column=0, columnspan=5, sticky=W)

name = Label(frame_left, text="Name : ", font=("Arial Black", 15), bg="#8B0000")
name.grid(row=1, column=0, pady=5, sticky=W)

ur_name = Entry(frame_left, font=("Arial Black", 15))
ur_name.grid(row=1, column=1, pady=5, columnspan=4)

age = Label(frame_left, text="Age :", font=("Arial Black", 15), bg="#8B0000")
age.grid(row=2, column=0, sticky=W, pady=10)
age_var = StringVar()
age_list = []
for i in range(18,81):
    age_list.append(i)

age_combo = ttk.Combobox(frame_left, values=age_list, textvariable=age_var, font=("Arial Black", 15), state="readonly")
age_combo.grid(row=2, column=1, columnspan=4)

movie = Label(frame_left, text="Pick Movies :", font=("Arial Black", 15), bg="#8B0000")
movie.grid(row=3, column=0)

ur_mov_var = StringVar()
ur_mov_list = ["Avatar", "Avengers", "Black Adam", "Chappie", "Fast X", "Transformer"]
ur_movie = ttk.Combobox(frame_left, font=("Arial Black", 15), state="readonly", values=ur_mov_list, textvariable=ur_mov_var)
ur_movie.grid(row=3, column=1, padx=10, columnspan=4)

time_ = Label(frame_left, text="Time : ", font=("Arial Black", 15), bg="#8B0000")
time_.grid(row=4, column=0, pady=5, sticky=W)
time_list = ["8:00AM - 10:30AM", "11:00AM - 1:30PM", "2:00PM - 4:30PM", "5:00PM - 7:30PM"]
time_var = StringVar()

time1_ = Radiobutton(frame_left, text=time_list[0], variable=time_var, value=time_list[0], font=("Arial Black", 15), bg="#8B0000")
time1_.grid(row=4, column=1, sticky=W, pady=5, padx=10, columnspan=4)

time2_ = Radiobutton(frame_left, text=time_list[1], variable=time_var, value=time_list[1], font=("Arial Black", 15), bg="#8B0000")
time2_.grid(row=5, column=1, sticky=W, pady=5, padx=10, columnspan=4)

time3_ = Radiobutton(frame_left, text=time_list[2], variable=time_var, value=time_list[2], font=("Arial Black", 15), bg="#8B0000")
time3_.grid(row=6, column=1, sticky=W, pady=5, padx=10, columnspan=4)

time4_ = Radiobutton(frame_left, text=time_list[3], variable=time_var, value=time_list[3], font=("Arial Black", 15), bg="#8B0000")
time4_.grid(row=7, column=1, sticky=W, pady=5, padx=10, columnspan=4)

seat_label = Label(frame_left, text="Pick Seat : ", font=("Arial Black",15), bg="#8B0000")
seat_label.grid(row=8, column=0, sticky=W)

seat_var = StringVar()
seat_list_a = []
for a in range(1,12):
    seat_list_a.append("A"+str(a))
    
seat_list_b = []
for b in range(1,12):
    seat_list_b.append("B"+str(b))
    
seat_list_c = []
for c in range(1,12):
    seat_list_c.append("C"+str(c))
    
seat_list_d = []
for d in range(1,12):
    seat_list_d.append("D"+str(d))
    
seat_list_e = []
for e in range(1,12):
    seat_list_e.append("E"+str(e))
    
seat_list_f = []
for f in range(1,12):
    seat_list_f.append("F"+str(f))
    
seat_combo_a = ttk.Combobox(frame_left, font=("Arial Black", 15), state="readonly",width=3, values=seat_list_a, textvariable=seat_var)
seat_combo_a.grid(row=8, column=1, columnspan=3, pady=5)

seat_combo_b = ttk.Combobox(frame_left, font=("Arial Black", 15), state="readonly",width=3, values=seat_list_b, textvariable=seat_var)
seat_combo_b.grid(row=8, column=2, columnspan=3, pady=5)

seat_combo_c = ttk.Combobox(frame_left, font=("Arial Black", 15), state="readonly",width=3, values=seat_list_c, textvariable=seat_var)
seat_combo_c.grid(row=8, column=4, columnspan=3, pady=5)

seat_combo_d = ttk.Combobox(frame_left, font=("Arial Black", 15), state="readonly",width=3, values=seat_list_d, textvariable=seat_var)
seat_combo_d.grid(row=9, column=1, columnspan=3, pady=5)

seat_combo_e = ttk.Combobox(frame_left, font=("Arial Black", 15), state="readonly",width=3, values=seat_list_e, textvariable=seat_var)
seat_combo_e.grid(row=9, column=2, columnspan=3, pady=5)

seat_combo_f = ttk.Combobox(frame_left, font=("Arial Black", 15), state="readonly",width=3, values=seat_list_f, textvariable=seat_var)
seat_combo_f.grid(row=9, column=4, columnspan=3, pady=5)


show_btn = Button(frame_left, text="Reserved", font=("Arial Black", 15), width=15, command=lambda:reserved_btn(), relief="raised", borderwidth=5)
show_btn.grid(row=11, column=0, pady=5, columnspan=5, sticky=W)

exit_btn = Button(frame_left, text="Exit", font=("Arial Black", 15), width=15, command=lambda:exit_btn_(), relief="raised", borderwidth=5)
exit_btn.grid(row=11, column=2, pady=5,sticky=E, columnspan=5)

done = Button(frame_left, text="Done", font=("Arial Black", 15), width=30, command=lambda:done_btn(), relief="raised", borderwidth=5)
done.grid(row=10, column=0, columnspan=5)

movie1 = "Avatar1.png"
movie2 = "Avengers.png"
movie3 = "Black Adam1.png"
movie4 = "Chappie1.png"
movie5 = "Fast X1.png"
movie6 = "Transformers1.png"
seat = "Seats.png"


img = ImageTk.PhotoImage(Image.open(movie1))
panel = tk.Label(frame_right, image=img, bg="Black")
panel.grid(row=0, column=0)

img1 = ImageTk.PhotoImage(Image.open(movie2))
panel1 = tk.Label(frame_right, image=img1, bg="Black")
panel1.grid(row=0, column=1)

img2 = ImageTk.PhotoImage(Image.open(movie3))
panel2 = tk.Label(frame_right, image=img2, bg="Black")
panel2.grid(row=0, column=2)

img3 = ImageTk.PhotoImage(Image.open(movie4))
panel3 = tk.Label(frame_right, image=img3, bg="Black")
panel3.grid(row=0, column=3)

img4 = ImageTk.PhotoImage(Image.open(movie5))
panel4 = tk.Label(frame_right, image=img4, bg="Black")
panel4.grid(row=0, column=4)

img5 = ImageTk.PhotoImage(Image.open(movie6))
panel5 = tk.Label(frame_right, image=img5, bg="Black")
panel5.grid(row=0, column=5)

img6 = ImageTk.PhotoImage(Image.open(seat))
panel6 = tk.Label(frame_right, image=img6, bg="Black")
panel6.grid(row=1, column=0,columnspan=7)

time_label = tk.Label(frame_left, font=("Arial Black", 15), bg="#8B0000")
time_label.grid(row=0, column=3, columnspan=4, padx=10)

date = dt.datetime.now()
label = Label(frame_left, text=f"{date:%A, %B %d, %Y}", font=("Arial Black", 15), bg="#8B0000")
label.grid(row=0, column=3, columnspan=4, sticky=N, padx=10)

update_clock()

root.mainloop()
